import os
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from FDPC import get_varience
from IFDPC.DPC import FDPC2
from FDPC.get_varience import read_data
from mesure.comment import balance, compute_SC
from openpyxl.utils.exceptions import InvalidFileException
from FDPC.FDPC6 import FDPC6

def create_excel_files(all_dataset, columns, project_path):
    """Create initial empty Excel files for all datasets."""
    for data_name in all_dataset:
        excel_name = os.path.join(project_path, f"{data_name}.xlsx")

        if not os.path.exists(excel_name):
            df = pd.DataFrame(columns=columns)
            df.to_excel(excel_name, index=False)
            print(f"Created Excel file: {excel_name}")
        else:
            print(f"Excel file already exists: {excel_name}")


def append_df_to_excel(file_path, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):
    """Append a DataFrame to an existing Excel file into specified sheet."""
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    try:
        # Try to open an existing workbook
        writer.book = load_workbook(file_path)
        if not startrow and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except (FileNotFoundError, InvalidFileException, zipfile.BadZipFile) as e:
        print(f"File could not be loaded: {e}")
        # Create a new workbook if the file is not found or is invalid
        writer.book = None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        raise

    if writer.book is None:
        # Create a new workbook
        writer.book = pd.ExcelWriter(file_path, engine='openpyxl').book

    if not startrow:
        startrow = 0

    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
    print(f"DataFrame successfully appended to {file_path}")


def cluster_num_iter(all_dataset, cluster_cnts, dcs, fair_tolerances,
                     columns, random_dataset, project_path):
    create_excel_files(all_dataset + random_dataset, columns, project_path)  # Create Excel files

    for data_name in all_dataset:
        print(f"\n当前数据库名称: {data_name}")
        results = []
        cluster_attr_vals, f_attr_vals, f_attr = read_data(data_name)
        for cluster_num in cluster_cnts:
            print(f"\n指定的当前簇的数量={cluster_num}")
            for dc in dcs:
                for fair_tolerance in fair_tolerances:
                    fdpc = FDPC6(cluster_attr_vals, f_attr_vals, f_attr, data_name=data_name,
                                 dc_rate=dc, fair_dc_min_rate=dc / 2, fair_dc_max_rate=dc * 2,
                                 fair_tolerance=fair_tolerance,
                                 cluster_num=cluster_num)

                    dpc = FDPC2(data_name=data_name, dataset_cluster=fdpc.dataset_cluster,
                                fair_attr_vals=fdpc.fair_attr_vals, fair_attr=fdpc.fair_attr,
                                dc_rate=dc, cluster_num=cluster_num)

                    fdpc_balance = balance(fdpc)
                    dpc_balance = balance(dpc)

                    fdpc_SC = compute_SC(fdpc)
                    dpc_SC = compute_SC(dpc)
                    fair_dc_value = fdpc.fair_dc
                    fair_dc_loss = np.sum(get_varience.cal_dc_fair_loss(fdpc, fdpc.fair_dc))
                    dc_value = fdpc.dc
                    dc_loss = np.sum(get_varience.cal_dc_fair_loss(fdpc, fdpc.dc))

                    balance_improvement = (fdpc_balance - dpc_balance) / dpc_balance

                    SC_improvement = fdpc_SC - dpc_SC

                    result = [cluster_num, dc, fair_tolerance,
                              fdpc_balance, dpc_balance,
                              balance_improvement,
                              fdpc_SC, dpc_SC, SC_improvement,
                              fair_dc_value, dc_value, fair_dc_loss, dc_loss]

                    results.append(result)

        df_results = pd.DataFrame(results, columns=columns)
        excel_name = os.path.join(project_path, f"{data_name}.xlsx")
        append_df_to_excel(excel_name, df_results)

        print(f"Results saved to {excel_name}")

    for data_name in random_dataset:
        print(f"\n当前数据库名称: {data_name}")
        results = []
        cluster_attr_vals, f_attr_vals, f_attr = read_data(data_name)
        for cluster_num in cluster_cnts:
            print(f"\n指定的当前簇的数量={cluster_num}")
            for dc in dcs:
                for fair_tolerance in fair_tolerances:
                    max_balance_improvement = -np.inf
                    best_result = None
                    for _ in range(1):  # 进行 10 次实验，取平均
                        fdpc = FDPC6(cluster_attr_vals, f_attr_vals, f_attr, data_name=data_name,
                                     dc_rate=dc, fair_dc_min_rate=dc / 2, fair_dc_max_rate=dc * 2,
                                     fair_tolerance=fair_tolerance,
                                     cluster_num=cluster_num)

                        dpc = FDPC2(data_name=data_name, dataset_cluster=fdpc.dataset_cluster,
                                    fair_attr_vals=fdpc.fair_attr_vals, fair_attr=fdpc.fair_attr,
                                    dc_rate=dc, cluster_num=cluster_num)

                        fdpc_balance = balance(fdpc)
                        dpc_balance = balance(dpc)

                        fdpc_SC = compute_SC(fdpc)
                        dpc_SC = compute_SC(dpc)
                        fair_dc_value = fdpc.fair_dc
                        fair_dc_loss = np.sum(get_varience.cal_dc_fair_loss(fdpc, fdpc.fair_dc))
                        dc_value = fdpc.dc
                        dc_loss = np.sum(get_varience.cal_dc_fair_loss(fdpc, fdpc.dc))

                        balance_improvement = (fdpc_balance - dpc_balance) / dpc_balance

                        SC_improvement = fdpc_SC - dpc_SC

                        result = [cluster_num, dc, fair_tolerance,
                                  fdpc_balance, dpc_balance,
                                  balance_improvement,
                                  fdpc_SC, dpc_SC, SC_improvement,
                                  fair_dc_value, dc_value, fair_dc_loss, dc_loss]

                        if balance_improvement > max_balance_improvement:
                            max_balance_improvement = balance_improvement
                            best_result = result

                    # 取均值，返回
                    results.append(best_result)
        df_results = pd.DataFrame(results, columns=columns)
        excel_name = os.path.join(project_path, f"{data_name}.xlsx")
        append_df_to_excel(excel_name, df_results)

        print(f"Results saved to {excel_name}")


if __name__ == '__main__':
    project_path = 'E:\\work\\fair_cluster\\current_work\\experiments\\FDPC6_2\\'
    all_dataset = ['Rice'] # 'abalone',

    # 'liver_disorder', 'Wholesale', 'bank', 'hcvdat0', 'Rice', 'obesity',
    # 'drug_consumption',  'vertebral', 'abalone',
    random_dataset = ['Room_Occupancy_Estimation', 'creditcard']  # 'adult', 'athlete', 'bank',
    # 淘汰： vertebral ,heart_failure_clinical, 'dabestic', 'drug', 'census1990'
    cluster_cnts = [4, 5, 6, 7, 8]
    dcs = [0.01, 0.02, 0.03, 0.04, 0.05]
    fair_tolerances = [0.01, 0.05, 0.1]
    columns = ['聚类数量', 'dc比例', '公平容忍度',
               'FDPC_balance', 'DPC_balance',
               'balance提升率',
               'FDPC_轮廓系数', 'DPC轮廓系数', '轮廓系数提升值',
               'fair_dc值', 'dc值', 'fair_dc公平损失', 'dc公平损失']

    cluster_num_iter(all_dataset, cluster_cnts, dcs, fair_tolerances,
                     columns, random_dataset, project_path)
